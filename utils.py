from openpyxl import load_workbook
from typing import Dict, List
import re

def clean_value(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace("$", "").replace("%", "")
        val = re.sub(r"[^\d\.\-]", "", val)
    try:
        return float(val)
    except:
        return None

def load_excel_sections(path: str) -> Dict[str, List[List[str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    sections = {}
    current_section = None
    section_rows = []

    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue  # skip blank rows

        # Heuristic: section header if the first non-empty cell is all caps or ends with '='
        first_cell = str(row[0]).strip() if row[0] else ''
        if (first_cell.isupper() or first_cell.endswith('=')) and sum(bool(cell) for cell in row) < 4:
            if current_section and section_rows:
                sections[current_section] = section_rows
                section_rows = []
            current_section = first_cell
        elif current_section:
            section_rows.append([str(cell).strip() if cell else '' for cell in row])

    if current_section and section_rows:
        sections[current_section] = section_rows

    return sections

def get_row_names(section_data: List[List[str]]) -> List[str]:
    return [row[0] for row in section_data if row and row[0]]

def get_row_sum(section_data: List[List[str]], row_name: str) -> float:
    for row in section_data:
        if row and row[0] == row_name:
            return sum(filter(None, (clean_value(cell) for cell in row[1:])))
    raise ValueError(f"Row '{row_name}' not found")
